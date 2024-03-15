import openpyxl
import PySimpleGUI as sg
from openpyxl import Workbook, load_workbook
import webbrowser
import random
from playsound import playsound
import os
def play_system_sound(sound_name):
    os.system("afplay /System/Library/Sounds/{}.aiff".format(sound_name))

book = load_workbook('data/IFAircraftData.xlsx')
sheet = book['main']
dataVerisonNumber = sheet['C1'].value


Checklist = load_workbook('data/Checklist.xlsx')
options = load_workbook('data/options.xlsx')

desired_altitudes = 0


def round_nrst_10(number):
    return round(number / 10) * 10

def InputLoadtoDataRow(acLoad):
    acLoad = int(acLoad)
    if acLoad == 23 or acLoad == 24 or acLoad == 25 or acLoad == 26 or acLoad == 27:
        row = 10
    elif acLoad == 73 or acLoad == 74 or acLoad == 75 or acLoad == 76 or acLoad == 77:
        row = 4
    else:
        acLoad = round_nrst_10(acLoad)
        if acLoad == 0 or acLoad == 10:
            row = 11
        elif acLoad == 20:
            row = 10
        elif acLoad == 30:
            row = 9
        elif acLoad == 40:
            row = 8
        elif acLoad == 50:
            row = 7
        elif acLoad == 60:
            row = 6
        elif acLoad == 70:
            row = 5
        elif acLoad == 80:
            row = 4
        elif acLoad == 90:
            row = 3
        elif acLoad == 100:
            row = 2

    return str(row)

def getDepatureData(row):
    DepPower = sheet['B' + row].value
    DepFlaps = sheet['C' + row].value
    DepRotate = sheet['D' + row].value
    DepAirBy = sheet['E' + row].value
    Author = sheet['D17'].value
    link = sheet['A17'].value
    return DepPower, DepFlaps, DepRotate, DepAirBy, Author, link

def getArrivalData(row):
    LdgFlaps = sheet['F'+ str(row)].value
    LdgApprSpd  = sheet['G'+ str(row)].value
    LdgFlareSpd = sheet['H'+ str(row)].value
    FlapsSpd = sheet['A14'].value
    Author = sheet['D17'].value
    link = sheet['A17'].value
    return LdgFlaps, LdgApprSpd, LdgFlareSpd, FlapsSpd, Author, link

def getFuelBurnData(row):
    Even = sheet['I'+ str(row)].value
    Odd = sheet['J'+ str(row)].value
    MedBurn = sheet['K'+ str(row)].value
    RecWest = sheet['E14'].value
    RecEast = sheet['F14'].value
    Author = sheet['D18'].value
    link1 = sheet['A18'].value
    link2 = sheet['A19'].value
    return Even, Odd, MedBurn, RecWest, RecEast, Author, link1, link2

def getOtherData(row):
    Ceiling = sheet['B14'].value
    Cruise = sheet['C14'].value
    MMO = sheet['D14'].value
    Range = sheet['A13'].value
    return Ceiling, Cruise, MMO, Range

def CL350(oat):
    if oat == -30:
        return "80% = 81.4% N1"
    elif oat == -25:
        return "81% = 82.3% N1"
    elif oat == -20:
        return "82% = 83.2% N1"
    elif oat == -15:
        return "83% = 84.0% N1"
    elif oat == -10:
        return "84% = 84.7% N1"
    elif oat == -5:
        return "85% = 85.5% N1"
    elif oat == 0:
        return "87% = 86.3% N1"
    elif oat == 5:
        return "87% = 87.0% N1"
    elif oat == 10:
        return "88% = 87.8% N1"
    elif oat == 15:
        return "90% = 88.7% N1"
    elif oat == 20:
        return "91% = 89.5% N1"
    elif oat == 25:
        return "92% = 90.3% N1"
    elif oat == 30 or oat == 35:
        return "93% = 91.1% N1"
    elif oat == 40:
        return "90% = 88.7% N1"
    else:
        return "Invalid input"

def getChecklistItem(page, itemNumber):
    sheetName = Checklist.worksheets[page]
    output = [sheetName['A' + str(itemNumber)].value, sheetName['B' + str(itemNumber)].value, sheetName['C' + str(itemNumber)].value, sheetName['D' + str(itemNumber)].value]
    if output[0] is None:
        output[0] = ' '
    if output[1] is None:
        output[1] = ' '
    return (output)

def updateChecklistItem(page, itemNumber, newComplete):
    sheetName = Checklist.worksheets[page]
    sheetName['C' + str(itemNumber)] = newComplete

def collapse(layout, key):
    return sg.pin(sg.Column(layout, key=key))











dark_mode = {
    'BACKGROUND': '#1C1C1D', #color1
    'TEXT': '#FFFFFF', #color3
    'INPUT': '#404040', #color2
    'TEXT_INPUT': '#FFFFFF',
    'SCROLL': '#404040',
    'BUTTON': ('#FFFFFF', '#404040'),
    'PROGRESS': ('#FFFFFF', '#404040'),
    'BORDER': 1,
    'SLIDER_DEPTH': 0,
    'PROGRESS_DEPTH': 0,
}
sg.theme_add_new('darkMode', dark_mode)

light_mode = {
    'BACKGROUND': '#F9F9F9', #color1
    'TEXT': '#5c5c5c', #color2
    'INPUT': '#E0E1E2', #color3
    'TEXT_INPUT': '#5c5c5c',
    'SCROLL': '#E0E1E2',
    'BUTTON': ('#5c5c5c', '#E0E1E2'),
    'PROGRESS': ('#5c5c5c', '#E0E1E2'),
    'BORDER': 1,
    'SLIDER_DEPTH': 0,
    'PROGRESS_DEPTH': 0,
}
sg.theme_add_new('lightMode', light_mode)

mode = 'dark'
option = options.active
mode = option['B2'].value
hideTitle = option['B1'].value


if mode == 'light':
    color1 = '#F9F9F9'
    color2 = '#E0E1E2'
    color3 = '#5c5c5c'
    filename = 'img/flightSmartTitleAppVerisonLight.png'
    sg.theme('lightMode')
elif mode =='dark':
    color1 = '#1C1C1D'
    color2 = '#404040'
    color3 = '#FFFFFF'
    filename = 'img/flightSmartTitleAppVerison.png'
    sg.theme('darkMode')


def getLinks(row, column):
    if column == 1:
        linkTitles = option['B'+ str(row)].value
        link = option['C' + str(row)].value
    elif column == 2:
        linkTitles = option['D'+ str(row)].value
        link = option['E' + str(row)].value
    return linkTitles, link


linkTitles = option['B4'].value, option['B5'].value, option['B6'].value, option['B7'].value, option['B8'].value, option['B9'].value, option['D4'].value, option[
    'D5'].value, option['D6'].value, option['D7'].value, option['D8'].value, option['D9'].value
linkLinks = option['C4'].value, option['C5'].value, option['C6'].value, option['C7'].value, option['C8'].value, option['C9'].value, option['E4'].value, option[
    'E5'].value, option['E6'].value, option['E7'].value, option['E8'].value, option['E9'].value

title_row = [
        [sg.Image(filename=filename ,size=(940, 50))],
        [sg.HorizontalSeparator(color=color1)]
]



sectionNotepad = [
            [sg.Multiline(size=(60, 60), key='notepadData', no_scrollbar=False, pad=(0,0))]
        ]


sectionGSESim = [
            [sg.Button('Widebody', auto_size_button=False, size=(17,1), button_color=(color2,color3), key='-PSwide-', pad=(1,1)), sg.Button('Narrowbody', auto_size_button=False, size=(17,1),key='-PSnarrow-', pad=(1,1))],
            [sg.Button('Departure', auto_size_button=False, size=(17,1), button_color=(color2,color3), key='-PSdep-', pad=(1,1)), sg.Button('Arrival', auto_size_button=False, size=(17,1),key='-PSarr-', pad=(1,1))],
            [sg.Button("Load Passengers", auto_size_button=False, size=(15,1), pad=(0,1)), sg.ProgressBar(100, orientation='h', size=(20, 10), key='-BoardingProgress-')],
            [sg.Button("Load Cargo", auto_size_button=False ,size=(15,1), pad=(0,1)), sg.ProgressBar(100, orientation='h', size=(20, 10), key='-CargoProgress-')],
            [sg.Button("Load Catering", auto_size_button=False, size=(15,1), pad=(0,1)), sg.ProgressBar(100, orientation='h', size=(20, 10), key='-CateringProgress-')],
            [sg.Multiline("MSG LOG:\n",size=(40,20), pad=(0,3), disabled=True, key='GSESimOutput')]
        ]

sectionFuel = [

            [sg.Text("Fuel Remaining", pad=(0,0)), sg.Input('', key='-Fuel Remaining-', size=(8,1), pad=(0,0)), sg.Text('  lbs or kg', pad=(0,0)) ],
            [sg.Text("Fuel Flow", pad=(0,0)), sg.Input('', key='-Fuel Flow-', size=(8,1), pad=(0,0)), sg.Text('  lbs/hr or kg/hr', pad=(0,0))],
            [sg.Button('Compute', size=(40,1))],
            [sg.Multiline("", size=(40, 20), pad=(0, 3), disabled=True, key='ASCOutput')]
]


col1 = sg.Column([
    [sg.Frame('Input:', layout=[
        [sg.Text("Manufacturer:"), sg.Button('Airbus'), sg.Button('Boeing'), sg.Button('Bombardier'), sg.Button('Embraer'), sg.Button('McDonnell Douglas')],
        [sg.Text('Aircraft:'), sg.DropDown(['      '], key='selected_aircraft'), sg.Text('  Load:'), sg.InputText(key='load', size=(3, 1)), sg.Text('%'),
        sg.Text("Select OAT (°C):", key='askOAT', visible=False), sg.Slider(disable_number_display=True, range=(-30, 40), default_value=0, orientation="h", key="OATinput", size=(16, 20), resolution=5, visible=False), sg.Text('0',key='sliderDisplay', visible=False)],
        [sg.Text("Request:"), sg.Button('Departure Data'),  sg.Button('Fuel Burn Data'), sg.Button('Arrival Data'), sg.Button('Other Data'), ]], size=(600, 120) )],

    [sg.Frame('Output:', layout=[
        [sg.Multiline(size=(60, 16),  key='output', disabled=True, no_scrollbar=False)],
        [sg.Button('Data Source 1 (IFC)', k='DataLink1',), sg.Button('Data Source 2 (IFC)', k='DataLink2')],
        ], size=(295, 340)),

     sg.Frame('Notepad:', layout=[
[
    #      sg.Button('Notes', enable_events=True, button_color=(color2, color3), pad=(3, 0), k='-OPEN NOTEPAD-',               font=('Helvetica Neue', 14)),
          #sg.Button('GSE', enable_events=True, pad=(3, 0), k='-OPEN GSESim-', font=('Helvetica Neue', 14)),
         # sg.Button('Wind', enable_events=True, pad=(3, 0), k='-OPEN XWIND-', font=('Helvetica Neue', 14)),
          #sg.Button('Time', enable_events=True, pad=(3, 0), k='-OPEN TIME-', font=('Helvetica Neue', 14)),
          #sg.Button('Fuel Burn', enable_events=True, pad=(3, 0), k='-OPEN FUEL-', font=('Helvetica Neue', 14)),

          ],
         #[sg.Combo(['Notepad', 'GSEsim', 'Wind Components', 'Time', 'Fuel', 'Flight Report'], default_value='Notepad', enable_events=True, size=(35,0), pad=(3, 3), k='-DROPDOWN-',)],
         [collapse(sectionNotepad, '-Notepad-')],
         [collapse(sectionGSESim, '-GSESim-')],
         [collapse(sectionFuel, '-Fuel-')],

     ], size=(295, 340))],

    [sg.Frame('App:', layout=[
        [sg.Button("Credits"), sg.Button("App Details ↗"), sg.Button("Dark / Light Mode", key="darkLightSwitcher"),
         sg.Button("Quit"), sg.Text('Verison 2' + sheet['C1'].value)],
    ], size=(600, 60))],

], expand_y=True)

#[sg.Button("Credits"), sg.Button("App Details ↗"), sg.Button("Dark / Light Mode", key="darkLightSwitcher"),
# sg.Button("Quit"), sg.Text("Quit and restart app to see update!", visible=False, key='quitRequest')],


if option['B3'].value == 'on':
    OKbuttonWidth = 14
    showMute = True
else:
    OKbuttonWidth = 19
    showMute = False

col2 = sg.Column([
        [sg.Frame('Checklist:', layout=[
        [sg.Text(key='checklist-Header', text_color=color3, background_color=color2, size=(65, 1), pad=(0,0), justification='center', enable_events=True, visible=False,)],
        [sg.Text(key='checklist-line1a', text_color=color3, background_color=color1, size=(16, 1), pad=(0,0), enable_events=True), sg.Text(key='checklist-line1b', text_color=color3, background_color=color1, expand_x=True, size=(16, 1), pad=(0,0), justification='right',enable_events=True)],
        [sg.Text(key='checklist-line2a', text_color=color3, background_color=color1, size=(16, 1), pad=(0,0), enable_events=True), sg.Text(key='checklist-line2b', text_color=color3, background_color=color1, expand_x=True, size=(16, 1), pad=(0,0), justification='right',enable_events=True)],
        [sg.Text(key='checklist-line3a', text_color=color3, background_color=color1, size=(16, 1), pad=(0,0), enable_events=True), sg.Text(key='checklist-line3b', text_color=color3, background_color=color1, expand_x=True, size=(16, 1), pad=(0,0), justification='right',enable_events=True)],
        [sg.Text(key='checklist-line4a', text_color=color3, background_color=color1, size=(16, 1), pad=(0,0), enable_events=True), sg.Text(key='checklist-line4b', text_color=color3, background_color=color1, expand_x=True, size=(16, 1), pad=(0,0), justification='right',enable_events=True)],
        [sg.Text(key='checklist-line5a', text_color=color3, background_color=color1, size=(16, 1), pad=(0,0), enable_events=True), sg.Text(key='checklist-line5b', text_color=color3, background_color=color1, expand_x=True, size=(16, 1), pad=(0,0), justification='right',enable_events=True)],
        [sg.Text(key='checklist-line6a', text_color=color3, background_color=color1, size=(16, 1), pad=(0,0), enable_events=True), sg.Text(key='checklist-line6b', text_color=color3, background_color=color1, expand_x=True, size=(16, 1), pad=(0,0), justification='right',enable_events=True)],
        [sg.Text(key='checklist-line7a', text_color=color3, background_color=color1, size=(16, 1), pad=(0,0), enable_events=True), sg.Text(key='checklist-line7b', text_color=color3, background_color=color1, expand_x=True, size=(16, 1), pad=(0,0), justification='right',enable_events=True)],
        [sg.Text(key='checklist-line8a', text_color=color3, background_color=color1, size=(16, 1), pad=(0,0), enable_events=True), sg.Text(key='checklist-line8b', text_color=color3, background_color=color1, expand_x=True, size=(16, 1), pad=(0,0), justification='right',enable_events=True)],
        [sg.Text(key='checklist-line9a', text_color=color3, background_color=color1, size=(16, 1), pad=(0,0), enable_events=True), sg.Text(key='checklist-line9b', text_color=color3, background_color=color1, expand_x=True, size=(16, 1), pad=(0,0), justification='right',enable_events=True)],
        [sg.Text(key='checklist-line10a', text_color=color3, background_color=color1, size=(16, 1), pad=(0,0), enable_events=True), sg.Text(key='checklist-line10b', text_color=color3, background_color=color1, expand_x=True, size=(16, 1), pad=(0,0), justification='right',enable_events=True)],
        [sg.HorizontalSeparator(color=color1)],
        [sg.Button("BACK", visible=False),sg.Button("OK", size=(OKbuttonWidth,1), auto_size_button=False, visible=False), sg.Button('🔊', size=(3,1), key='muteSwitch', auto_size_button=False, visible=False), sg.Button("NEXT", visible=False)],
        [sg.Button("Start", size=(35,1), auto_size_button=False, visible=True)]
        ], size=(295, 300))],
        [sg.Frame('Hyperlinks: ↗', layout=[
        [sg.Button(linkTitles[0], size=(15,1), auto_size_button=False), sg.Text(''), sg.Button(linkTitles[6], size=(15,1), auto_size_button=False,)],
        [sg.Button(linkTitles[1], size=(15,1), auto_size_button=False,), sg.Text(''), sg.Button(linkTitles[7], size=(15,1), auto_size_button=False,)],
        [sg.Button(linkTitles[2], size=(15,1), auto_size_button=False,), sg.Text(''), sg.Button(linkTitles[8], size=(15,1), auto_size_button=False,)],
        [sg.Button(linkTitles[3], size=(15,1), auto_size_button=False,), sg.Text(''), sg.Button(linkTitles[9], size=(15,1), auto_size_button=False,)],
        [sg.Button(linkTitles[4], size=(15,1), auto_size_button=False,), sg.Text(''), sg.Button(linkTitles[10], size=(15,1), auto_size_button=False,)],
        [sg.Button(linkTitles[5], size=(15,1), auto_size_button=False,), sg.Text(''), sg.Button(linkTitles[11], size=(15,1), auto_size_button=False,)],



        ], size=(295, 240))],
], expand_y=True)



if hideTitle == 'yes':
    windowHeight = 555
    title = 'flightSmart'
    layout = [
        [col1, col2],
    ]
else:
    windowHeight = 620
    title = ' '
    layout = [
        title_row,
        [col1, col2],


    ]



window = sg.Window("flightSmart", layout, font=('Helvetica Neue', 14), size=(955, windowHeight), resizable=True, finalize=True)  # size: width, height

sheetNumber = -1

currentPage = -1
currentItem = 0
itemNumber = 0


window.bind('<Right>', 'NEXT')
window.bind('<Down>', 'OK')
window.bind('<Space>', 'OK')
window.bind('<Left>', 'BACK')



PaxTimer_running = False
PaxTimer_value = 0
CargoTimer_running = False
CargoTimer_value = 0
CateringTimer_running = False
CateringTimer_value = 0

PStype = 'wide'
PSdeparr = 'dep'


muted = False
pressedOnce = False

currentLink = 'Empty'
currentLink2 = 'Empty'

while True:
    event, value = window.read(timeout=1000)


    if event == "Quit" or event == sg.WIN_CLOSED:
        break

    if event == 'DataLink1':
        if currentLink != 'Empty':
            try:
                webbrowser.open(currentLink)
            except:
                print(" ")

    if event == 'DataLink2':

        if currentLink != 'Empty':
            try:
                webbrowser.open(currentLink2)
            except:
                print("")


    if event == 'darkLightSwitcher':
        if pressedOnce == False:
            if mode == 'dark':
                pressedOnce = True
                option['B2'].value = 'light'
                options.save("data/options.xlsx")
                sg.Popup("Quit and restart app for light mode!")
            elif mode == 'light':
                pressedOnce = True
                option['B2'].value = 'dark'
                options.save("data/options.xlsx")
                sg.Popup("Quit and restart app for dark mode")



    window['sliderDisplay'].update(value['OATinput'])



    if event == 'NEXT' or event =='Start':
        window['Start'].update(visible=False)
        window['checklist-Header'].update(visible=True)
        window['BACK'].update(visible=True)
        window['OK'].update(visible=True)
        window['NEXT'].update(visible=True)
        window['muteSwitch'].update(visible=showMute)

        try:


            currentPage += 1

            sheetName = Checklist.worksheets[currentPage]
            window['checklist-Header'].update(sheetName.title)
            if muted != True:
                try:
                    playsound('data/checklistAudio/' + sheetName.title + '.mp3', False)
                except:
                    print('')

            while itemNumber <= 9:
                itemNumber += 1
                itemData = getChecklistItem(currentPage, itemNumber)
                if itemData[2] == 'uncompleted':
                    window['checklist-line' + str(itemNumber) +'a'].update(itemData[0], text_color=color3, background_color=color1)
                    window['checklist-line' + str(itemNumber) + 'b'].update(itemData[1], text_color=color3, background_color=color1)
                if itemData[2] == 'completed':
                    window['checklist-line' + str(itemNumber) +'a'].update(itemData[0], text_color='green', background_color=color1)
                    window['checklist-line' + str(itemNumber) + 'b'].update(itemData[1], text_color='green',background_color=color1)

            itemNumber = 0
            currentItem = 0
        except:
            currentPage -= 1
    elif event == 'muteSwitch':
        if muted == False:
            muted = True
            window['muteSwitch'].update('🔇')
        else:
            muted = False
            window['muteSwitch'].update('🔊')
    elif event == 'OK':
        currentItem += 1
        previousItem = currentItem - 1
        if currentItem <= 10:
            itemData = getChecklistItem(currentPage, currentItem)
            if muted != True:
                try:
                    playsound('data/checklistAudio/' + itemData[0] + '.mp3', False)
                except:
                    print('')
            if itemData[1] == ' ':
                window['checklist-line' + str(currentItem) + 'a'].update(itemData[0], text_color='green', background_color=color1)
                window['checklist-line' + str(currentItem) + 'b'].update(itemData[1], text_color='green', background_color=color1)
                previousItemData = getChecklistItem(currentPage, previousItem)
                window['checklist-line' + str(previousItem) + 'a'].update(previousItemData[0], text_color='green', background_color=color1)
                window['checklist-line' + str(previousItem) + 'b'].update(previousItemData[1], text_color='green', background_color=color1)
                updateChecklistItem(currentPage, currentItem, 'completed')
                updateChecklistItem(currentPage, previousItem, 'completed')
            else:
                nextItem = currentItem + 1
                window['checklist-line' + str(currentItem) + 'a'].update(itemData[0], text_color=color3, background_color='green')
                window['checklist-line' + str(currentItem) + 'b'].update(itemData[1], text_color=color3, background_color='green')
                updateChecklistItem(currentPage, currentItem, 'uncompleted')
                if previousItem != 0:
                    previousItemData = getChecklistItem(currentPage, previousItem)
                    window['checklist-line' + str(previousItem) + 'a'].update(previousItemData[0], text_color='green', background_color=color1)
                    window['checklist-line' + str(previousItem) + 'b'].update(previousItemData[1], text_color='green', background_color=color1)
                    updateChecklistItem(currentPage, previousItem, 'completed')
        elif currentItem == 11:
            currentItem = 10
            window['checklist-line' + str(currentItem) + 'a'].update(itemData[0], text_color='green', background_color=color1)
            window['checklist-line' + str(currentItem) + 'b'].update(itemData[1], text_color='green', background_color=color1)
            updateChecklistItem(currentPage, currentItem, 'completed')

    elif event == 'BACK':
        try:
            currentPage -= 1

            sheetName = Checklist.worksheets[currentPage]
            window['checklist-Header'].update(sheetName.title)

            while itemNumber <= 9:
                itemNumber += 1
                itemData = getChecklistItem(currentPage, itemNumber)
                if itemData[2] == 'uncompleted':
                    window['checklist-line' + str(itemNumber) +'a'].update(itemData[0], text_color=color3, background_color=color1)
                    window['checklist-line' + str(itemNumber) + 'b'].update(itemData[1], text_color=color3, background_color=color1)
                if itemData[2] == 'completed':
                    window['checklist-line' + str(itemNumber) +'a'].update(itemData[0], text_color='green', background_color=color1)
                    window['checklist-line' + str(itemNumber) + 'b'].update(itemData[1], text_color='green',background_color=color1)
            itemNumber = 0
            currentItem = 0
        except:
            currentPage += 1

    elif event == 'Airbus':
        window['Airbus'].update(button_color = (color2,color3 )); window['Boeing'].update(button_color=(color3, color2)); window['Bombardier'].update(button_color=(color3, color2)); window['Embraer'].update(button_color=(color3, color2)); window['McDonnell Douglas'].update(button_color=(color3, color2))
        new_choices = ['A220', 'A318', 'A319', 'A320', 'A321', 'A332', 'A333', 'A339', 'A346', 'A359', 'A388']
        window['selected_aircraft'].update(values=new_choices)

    elif event == 'Boeing':
        window['Boeing'].update(button_color = (color2,color3 )); window['Airbus'].update(button_color=(color3, color2)); window['Bombardier'].update(button_color=(color3, color2)); window['Embraer'].update(button_color=(color3, color2)); window['McDonnell Douglas'].update(button_color=(color3, color2))
        new_choices = ['B712', 'B737', 'B738', 'B739', 'B742', 'B744', 'B748', 'B752', 'B763', 'B772', 'B77L', 'B77W', 'B77F',
               'B788', 'B789', 'B78X']
        window['selected_aircraft'].update(values=new_choices)

    elif event == 'Bombardier':
        window['Bombardier'].update(button_color = (color2,color3 )); window['Boeing'].update(button_color=(color3, color2)); window['Airbus'].update(button_color=(color3, color2)); window['Embraer'].update(button_color=(color3, color2)); window['McDonnell Douglas'].update(button_color=(color3, color2))
        window['Bombardier'].update(button_color=(color2, color3))
        new_choices = ['CL350', 'CRJ2', 'CRJ7', 'CRJ9', 'CRJX', 'DH8D']
        window['selected_aircraft'].update(values=new_choices)

    elif event == 'Embraer':
        window['Embraer'].update(button_color = (color2,color3 )); window['Boeing'].update(button_color=(color3, color2)); window['Bombardier'].update(button_color=(color3, color2)); window['Airbus'].update(button_color=(color3, color2)); window['McDonnell Douglas'].update(button_color=(color3, color2))
        new_choices = ['E175', 'E190']
        window['selected_aircraft'].update(values=new_choices)

    elif event == 'McDonnell Douglas':
        window['McDonnell Douglas'].update(button_color = (color2,color3 )); window['Boeing'].update(button_color=(color3, color2)); window['Bombardier'].update(button_color=(color3, color2)); window['Embraer'].update(button_color=(color3, color2)); window['Airbus'].update(button_color=(color3, color2))
        new_choices = ['DC10', 'DC1F', 'MD11', 'MD1F']
        window['selected_aircraft'].update(values=new_choices)

    elif event == 'App Details ↗':
        webbrowser.open("https://community.infiniteflight.com/t/the-unofficial-infinite-aircraft-calculator-using-community-data/869648")

    elif event == 'Credits':
        window['Credits'].update(button_color = (color2,color3 )); window['Departure Data'].update(button_color=(color3, color2)); window['Arrival Data'].update(button_color=(color3, color2)); window['Fuel Burn Data'].update(button_color=(color3, color2)); window['Other Data'].update(button_color=(color3, color2));
        output_text = '\n'.join([
            f"Data From:"
            f"\n DeerCrusher, Kuba_Jaroszczy, AndrewWu, Jan, AviatorAlex, nicopizarro, ToasterStroodie, RickG, Topgottem",
            f"\n",
            f"App by"
            f"\n darkeyes",
            f"\n",
            f"\nThe data here isn't perfectly accurate. It is simply intended to offer a basic guidance for your flight. ",


        ])
        window['output'].update(output_text, text_color=color3)

    elif event == 'Departure Data':
        window['Departure Data'].update(button_color = (color2,color3 )); window['Credits'].update(button_color=(color3, color2)); window['Arrival Data'].update(button_color=(color3, color2)); window['Fuel Burn Data'].update(button_color=(color3, color2)); window['Other Data'].update(button_color=(color3, color2));
        try:
            acType = value["selected_aircraft"]
            sheet = book[acType]
            ac_load = value['load']
            row = InputLoadtoDataRow(ac_load)
            DepPower, DepFlaps, DepRotate, DepAirBy, Author, link = getDepatureData(row)
            if value["selected_aircraft"] == 'CL350':
                window['askOAT'].update(visible=True)
                window['OATinput'].update(visible=True)
                window['sliderDisplay'].update(visible=True)
                oat = value["OATinput"]
                DepPower = CL350(oat)
                output_text = '\n'.join([f"Flap Setting: {DepFlaps} ", f"\nPower: \n{DepPower} ", f"\nRotate: "
                            f"{DepRotate} ", f"Airborne By: {DepAirBy}", f"\nData By: {Author} "])
                currentLink = link
                currentLink2 = 'Empty'
            else:
                window['askOAT'].update(visible=False)
                window['OATinput'].update(visible=False)
                window['sliderDisplay'].update(visible=False)
                output_text = '\n'.join(
                    [f"Flap Setting: {DepFlaps} ", f"\nPower: {DepPower} ", f"\nRotate: {DepRotate} ",
                     f"Airborne By: {DepAirBy}", f"\nData By: {Author} "])
                currentLink = link
                currentLink2 = 'Empty'
            window['output'].update(output_text, text_color=color3)
        except KeyError:
            errorMessage = 'Select Aircraft Type!'
            window['output'].update(errorMessage, text_color='red')
        except ValueError:
            errorMessage = 'Enter Aircraft Load as an Integer!'
            window['output'].update(errorMessage, text_color='red')
        except:
            errorMessage = 'Error, try again'
            mycolor = 'red'
            window['output'].update(errorMessage, text_color='red')

    elif event == 'Arrival Data':
        window['Arrival Data'].update(button_color = (color2,color3 )); window['Credits'].update(button_color=(color3, color2)); window['Departure Data'].update(button_color=(color3, color2)); window['Fuel Burn Data'].update(button_color=(color3, color2)); window['Other Data'].update(button_color=(color3, color2));
        try:
            acType = value["selected_aircraft"]
            sheet = book[acType]
            ac_load = value['load']
            row = InputLoadtoDataRow(ac_load)
            ArrFlaps, ArrApprSpd, ArrFlareSpd, Flaps, Author, link = getArrivalData(row)
            output_text = '\n'.join([f"Flap Setting: {ArrFlaps} ", f"\nApproach Speed: {ArrApprSpd}", f"Flare Speed: {ArrFlareSpd} ", f"\nFlap Spds: \n{Flaps} ", f"\nData By: {Author} "])
            currentLink = link
            currentLink2 = 'Empty'
            window['output'].update(output_text, text_color=color3)
        except KeyError:
            errorMessage = 'Select Aircraft Type!'
            mycolor = 'red'
            window['output'].update(errorMessage, text_color='red')
        except ValueError:
            errorMessage = 'Enter Aircraft Load as an Integer!'
            mycolor = 'red'
            window['output'].update(errorMessage, text_color='red')
        except:
            errorMessage = 'Error, try again'
            mycolor = 'red'
            window['output'].update(errorMessage, text_color='red')

    elif event == 'Fuel Burn Data':
         window['Fuel Burn Data'].update(button_color = (color2,color3 )); window['Credits'].update(button_color=(color3, color2)); window['Departure Data'].update(button_color=(color3, color2)); window['Arrival Data'].update(button_color=(color3, color2)); window['Other Data'].update(button_color=(color3, color2));
         try:
            acType = value["selected_aircraft"]
            sheet = book[acType]
            ac_load = value['load']
            row = InputLoadtoDataRow(ac_load)
            Even, Odd, Med, RecWest, RecEast, author, link1, link2,  = getFuelBurnData(row)
            output_text = '\n'.join([f"West/Even Cruise Alt: {Even} ", f"East/Odd Cruise Alt: {Odd} ", f"High Fuel Burn: {Med} ",
                                      f"\n\nRecommend Flight Profile West/Even: {RecWest}", f"\nRecommend Flight Profile East/Odd: {RecEast}",  f"\nData By: {author} "])
            currentLink = link1
            currentLink2 = link2
            window['output'].update(output_text, text_color=color3)
         except KeyError:
            errorMessage = 'Select Aircraft Type!'
            mycolor = 'red'
            window['output'].update(errorMessage, text_color='red')
         except ValueError:
             errorMessage = 'Enter Aircraft Load as an Integer!'
             mycolor = 'red'
             window['output'].update(errorMessage, text_color='red')
         except:
             errorMessage = 'Error, try again'
             mycolor = 'red'
             window['output'].update(errorMessage, text_color='red')

    elif event == 'Other Data':
        window['Other Data'].update(button_color = (color2,color3 )); window['Credits'].update(button_color=(color3, color2)); window['Departure Data'].update(button_color=(color3, color2)); window['Arrival Data'].update(button_color=(color3, color2)); window['Fuel Burn Data'].update(button_color=(color3, color2));
        try:
            acType = value["selected_aircraft"]
            sheet = book[acType]
            ac_load = value['load']
            row = InputLoadtoDataRow(ac_load)
            Ceiling, Cruise, MMO, Range = getOtherData(row)
            output_text = '\n'.join([f"Ceiling: {Ceiling}", f"Normal Range: {Range}", f"\nCruise Spd: {Cruise}", f"MMO Spd: {MMO} \n",  f"Data by: {'Jan'} \n"])
            window['output'].update(output_text, text_color=color3)
            currentLink = 'https://www.helpathand.nl/janpolet/infinite-flight-aircraft-information/'
            currentLink2 = 'Empty'
        except KeyError:
            errorMessage = 'Select Aircraft Type!'
            mycolor = 'red'
            window['output'].update(errorMessage, text_color='red')
        except ValueError:
            errorMessage = 'Enter Aircraft Load as an Integer!'
            mycolor = 'red'
            window['output'].update(errorMessage, text_color='red')
        except:
            errorMessage = 'Error, try again'
            mycolor = 'red'
            window['output'].update(errorMessage, text_color='red')

    elif event == linkTitles[0]:
        webbrowser.open(linkLinks[0])
    elif event == linkTitles[1]:
        webbrowser.open(linkLinks[1])
    elif event == linkTitles[2]:
        webbrowser.open(linkLinks[2])
    elif event == linkTitles[3]:
        webbrowser.open(linkLinks[3])
    elif event == linkTitles[4]:
        webbrowser.open(linkLinks[4])
    elif event == linkTitles[5]:
        webbrowser.open(linkLinks[5])
    elif event == linkTitles[6]:
        webbrowser.open(linkLinks[6])
    elif event == linkTitles[7]:
        webbrowser.open(linkLinks[7])
    elif event == linkTitles[8]:
        webbrowser.open(linkLinks[8])
    elif event == linkTitles[9]:
        webbrowser.open(linkLinks[9])
    elif event == linkTitles[10]:
        webbrowser.open(linkLinks[10])
    elif event == linkTitles[11]:
        webbrowser.open(linkLinks[11])



options.save("options.xlsx")
window.close()
